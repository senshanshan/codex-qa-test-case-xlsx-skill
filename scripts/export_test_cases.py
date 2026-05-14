#!/usr/bin/env python
import argparse
import json
from collections import Counter
from copy import deepcopy
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_COLUMNS = [
    "用例编号",
    "模块",
    "用例标题",
    "前置条件",
    "测试步骤",
    "预期结果",
    "优先级",
    "类型",
    "备注",
]

DEFAULT_OUTPUT_DIR = Path.cwd()
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
INVALID_SHEET_CHARS = '[]:*?/\\\\'


def parse_args():
    parser = argparse.ArgumentParser(
        description="Export categorized test cases to an xlsx workbook."
    )
    parser.add_argument("--input-json", required=True, help="Path to structured test-case JSON.")
    parser.add_argument("--output-dir", help="Output directory for the workbook.")
    parser.add_argument("--file-name", help="Optional workbook file name.")
    return parser.parse_args()


def sanitize_filename_part(value):
    if not value:
        return ""
    invalid = '<>:"/\\|?*'
    sanitized = "".join("_" if ch in invalid else ch for ch in str(value).strip())
    return sanitized.strip(" ._")


def build_default_name(payload):
    today = datetime.now().strftime("%Y-%m-%d")
    system_name = sanitize_filename_part(payload.get("system_name", ""))
    module_name = sanitize_filename_part(payload.get("module_name", ""))

    if system_name and module_name:
        return f"{system_name}_{module_name}_测试用例_{today}.xlsx"
    if module_name:
        return f"{module_name}_测试用例_{today}.xlsx"
    return f"测试用例_{today}.xlsx"


def unique_path(directory, file_name):
    base = Path(directory)
    candidate = base / file_name
    if not candidate.exists():
        return candidate

    stem = candidate.stem
    suffix = candidate.suffix
    index = 1
    while True:
        next_candidate = base / f"{stem}_{index:02d}{suffix}"
        if not next_candidate.exists():
            return next_candidate
        index += 1


def normalize_payload(payload):
    normalized = deepcopy(payload)
    normalized.setdefault("system_name", "")
    normalized.setdefault("module_name", "")
    normalized.setdefault("input_summary", "")
    normalized.setdefault("generated_at", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    normalized.setdefault("output_format", "xlsx")
    normalized.setdefault("uncertainties", [])
    normalized.setdefault("extra_columns", [])
    normalized.setdefault("categories", [])
    return normalized


def build_workbook(payload):
    payload = normalize_payload(payload)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "说明"

    extra_columns = payload.get("extra_columns", [])
    all_columns = DEFAULT_COLUMNS + [c for c in extra_columns if c not in DEFAULT_COLUMNS]

    category_counts = Counter()
    total_cases = 0

    for category in payload["categories"]:
        category_name = str(category.get("name", "")).strip() or "未分类"
        cases = category.get("cases", [])
        category_counts[category_name] += len(cases)
        total_cases += len(cases)

        sheet = workbook.create_sheet(title=trim_sheet_name(category_name))
        write_case_sheet(sheet, all_columns, cases)

    write_summary_sheet(summary, payload, category_counts, total_cases)
    return workbook


def trim_sheet_name(name):
    cleaned = "".join("_" if ch in INVALID_SHEET_CHARS else ch for ch in name)
    return cleaned[:31] or "Sheet"


def write_case_sheet(sheet, columns, cases):
    sheet.append(columns)
    style_header_row(sheet, 1)
    sheet.freeze_panes = "A2"

    for case in cases:
        row = [case.get(column, "") for column in columns]
        sheet.append(row)

    apply_case_sheet_formatting(sheet)


def write_summary_sheet(sheet, payload, category_counts, total_cases):
    rows = [
        ("系统名", payload.get("system_name", "")),
        ("模块名", payload.get("module_name", "")),
        ("输入来源", payload.get("input_summary", "")),
        ("生成时间", payload.get("generated_at", "")),
        ("输出格式", payload.get("output_format", "xlsx")),
        ("用例总数", total_cases),
    ]

    for label, value in rows:
        sheet.append([label, value])

    sheet.append([])
    sheet.append(["分类", "数量"])
    style_header_row(sheet, sheet.max_row)

    for name, count in category_counts.items():
        sheet.append([name, count])

    sheet.append([])
    sheet.append(["待确认项"])
    style_header_row(sheet, sheet.max_row)

    uncertainties = payload.get("uncertainties", [])
    if uncertainties:
        for item in uncertainties:
            sheet.append([item])
    else:
        sheet.append(["无"])

    apply_summary_formatting(sheet)


def style_header_row(sheet, row_index):
    for cell in sheet[row_index]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP_ALIGNMENT


def apply_case_sheet_formatting(sheet):
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = WRAP_ALIGNMENT

    max_width = 40
    min_width = 12
    for column_cells in sheet.columns:
        letter = get_column_letter(column_cells[0].column)
        max_len = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[letter].width = max(min_width, min(max_width, max_len + 2))


def apply_summary_formatting(sheet):
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = WRAP_ALIGNMENT

    sheet.column_dimensions["A"].width = 18
    sheet.column_dimensions["B"].width = 60


def main():
    args = parse_args()
    input_path = Path(args.input_json)
    with input_path.open("r", encoding="utf-8") as fh:
        payload = json.load(fh)

    output_dir = Path(args.output_dir or payload.get("output_dir") or DEFAULT_OUTPUT_DIR)
    output_dir.mkdir(parents=True, exist_ok=True)

    file_name = args.file_name or payload.get("file_name") or build_default_name(payload)
    if not file_name.lower().endswith(".xlsx"):
        file_name = f"{file_name}.xlsx"

    workbook = build_workbook(payload)
    final_path = unique_path(output_dir, sanitize_filename_part(file_name[:-5]) + ".xlsx")
    workbook.save(final_path)
    print(str(final_path))


if __name__ == "__main__":
    main()
